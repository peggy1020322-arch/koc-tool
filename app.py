import os
import sys
import json
import re
import time
import hashlib
import requests
import urllib.parse
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, jsonify
import instaloader
import subprocess

warnings.filterwarnings("ignore")

app = Flask(__name__)

# ── Instagram session ──
_ig = None
_ig_username = None
_ig_password = None
_ig_pending = None

THREADS_STATE   = os.path.join(os.path.dirname(__file__), ".threads_state.json")
THREADS_COOKIES = os.path.join(os.path.dirname(__file__), ".threads_cookies.json")
WORKER          = os.path.join(os.path.dirname(__file__), "threads_worker.py")
IG_SESSION_FILE = os.path.join(os.path.dirname(__file__), ".ig_session")
IG_USERNAME_FILE= os.path.join(os.path.dirname(__file__), ".ig_username")


def _try_load_ig_session():
    """Server 啟動時嘗試從檔案還原 IG session"""
    global _ig, _ig_username
    if not os.path.exists(IG_SESSION_FILE) or not os.path.exists(IG_USERNAME_FILE):
        return
    try:
        username = open(IG_USERNAME_FILE).read().strip()
        loader = instaloader.Instaloader(
            download_pictures=False, download_videos=False,
            download_video_thumbnails=False, download_geotags=False,
            download_comments=False, save_metadata=False, quiet=True,
        )
        loader.load_session_from_file(username, IG_SESSION_FILE)
        _ig = loader
        _ig_username = username
    except Exception:
        pass


_try_load_ig_session()

# 簡易快取，避免同一搜尋重複觸發限流
_cache = {}
CACHE_TTL = 600  # 10 分鐘

def cache_get(key):
    if key in _cache:
        ts, val = _cache[key]
        if time.time() - ts < CACHE_TTL:
            return val
    return None

def cache_set(key, val):
    _cache[key] = (time.time(), val)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}
SKIP_HANDLES = {"tag", "explore", "login", "register", "about", "help", "privacy", "p"}

# Threads 行動版 API headers（Barcelona = Threads app）
THREADS_APP_HEADERS = {
    "User-Agent": (
        "Barcelona 289.0.0.77.109 Android (31; 480dpi; 1080x2400; "
        "OnePlus; GM1917; OnePlus7Pro; qcom; en_US; 289.0.0.77.109)"
    ),
    "X-IG-App-ID": "238260118697367",
    "Accept": "*/*",
    "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate",
}

# 爬蟲用 UA（模擬 FB crawler 拿到 OG tags）
CRAWLER_UA = "facebookexternalhit/1.1"


def parse_followers(desc):
    """從 og:description 解析粉絲數，例如 '1.5 萬位粉絲' → 15000"""
    m = re.search(r'([\d,.]+)\s*萬位粉絲', desc)
    if m:
        return int(float(m.group(1).replace(',', '')) * 10000)
    m = re.search(r'([\d,]+)\s*位粉絲', desc)
    if m:
        return int(m.group(1).replace(',', ''))
    return None


def parse_posts(desc):
    """從 og:description 解析貼文數，例如 '259 則串文' → 259"""
    m = re.search(r'([\d,]+)\s*則串文', desc)
    if m:
        return int(m.group(1).replace(',', ''))
    return None


def fetch_profile(handle):
    """抓取單一 Threads 帳號的個人資料（OG meta）"""
    username = handle.lstrip('@')
    cache_key = 'profile_' + username
    cached = cache_get(cache_key)
    if cached is not None:
        return cached

    url = f"https://www.threads.com/@{username}"
    try:
        resp = requests.get(url, headers={"User-Agent": CRAWLER_UA, "Accept-Language": "zh-TW,zh;q=0.9"}, timeout=8)
        soup = BeautifulSoup(resp.text, "html.parser")

        def og(prop):
            el = soup.find("meta", property=prop)
            return el.get("content", "") if el else ""

        title_raw = og("og:title")           # 有溫度的妝（@millet_makeup_） • Threads，暢所欲言
        desc_raw  = og("og:description")     # 1.5 萬位粉絲 • 259 則串文 • 透明感...
        image     = og("og:image")

        # 清理名稱（去掉後面的 • Threads…）
        display_name = re.split(r'[（•]', title_raw)[0].strip()

        # 解析 bio（粉絲數那行之後）
        bio = ""
        if "•" in desc_raw:
            parts = desc_raw.split("•", 2)
            bio = parts[2].strip() if len(parts) > 2 else ""
            bio = re.split(r'。查看', bio)[0].strip()

        followers = parse_followers(desc_raw)
        posts     = parse_posts(desc_raw)

        result = {
            "handle": handle,
            "display_name": display_name,
            "avatar": image,
            "bio": bio[:100] if bio else "",
            "followers": followers,
            "posts": posts,
        }
        cache_set(cache_key, result)
        return result
    except Exception:
        return {"handle": handle, "display_name": "", "avatar": "", "bio": "", "followers": None, "posts": None}

# 功能對應搜尋關鍵字組合
FEATURE_QUERIES = {
    "AI磨皮": ["AI磨皮 自拍", "磨皮濾鏡 美肌", "美顏相機 磨皮"],
    "美白瘦臉": ["美白 瘦臉 app", "瘦臉濾鏡 自拍", "美白修圖"],
    "換髮色": ["換髮色 app", "染髮 修圖 自拍", "髮色濾鏡"],
    "美顏相機": ["美顏相機", "美顏app 自拍", "美顏 修圖推薦"],
    "妝效濾鏡": ["妝效 濾鏡 自拍", "試妝 app", "彩妝濾鏡"],
    "自訂": [],  # 用戶自行輸入
}

# 標籤對應表（Tab 1 一般查找用）
CATEGORY_TAGS = {
    "美妝 彩妝": ["美妝", "彩妝", "美妝分享", "彩妝教學", "台灣美妝"],
    "保養 護膚": ["保養", "護膚", "保養推薦", "台灣保養", "skincare"],
    "美白 淡斑": ["美白", "淡斑", "保養推薦", "美白保養"],
    "自拍 修圖": ["自拍", "修圖", "美肌", "美顏", "selfie"],
    "穿搭 生活": ["穿搭", "生活", "ootd", "日常穿搭", "台灣穿搭"],
    "美髮 髮色": ["美髮", "髮色", "染髮", "台灣美髮"],
}


def ddg_search(query, max_results=20):
    """DuckDuckGo 搜尋，回傳 Threads 帳號＋貼文片段。有快取，10分鐘內同查詢不重複打。"""
    cache_key = hashlib.md5(query.encode()).hexdigest()
    cached = cache_get(cache_key)
    if cached is not None:
        return cached

    seen = set()
    results = []
    params = {"q": query, "kl": "tw-tzh"}
    resp = requests.get(
        "https://html.duckduckgo.com/html/",
        params=params, headers=HEADERS, timeout=14
    )

    if resp.status_code == 202:
        # 被限流，回傳空結果（不快取，讓下次可重試）
        return []

    soup = BeautifulSoup(resp.text, "html.parser")

    for a in soup.find_all("a"):
        href = a.get("href", "")
        if "threads.com" not in href:
            continue
        decoded = urllib.parse.unquote(href)
        m = re.search(r'threads\.com/@([\w.]+)', decoded)
        if not m:
            continue
        handle = m.group(1).lower()
        if handle in SKIP_HANDLES or handle in seen:
            continue
        seen.add(handle)

        # 抓貼文片段
        snippet = ""
        node = a
        for _ in range(6):
            node = node.parent
            if node is None:
                break
            txt = node.get_text(" ", strip=True)
            if len(txt) > 30:
                snippet = txt[:160]
                break

        # 找結果標題
        title = ""
        result_block = a.find_parent("div", class_=re.compile("result"))
        if result_block:
            t_el = result_block.select_one(".result__title, h2")
            if t_el:
                title = t_el.get_text(" ", strip=True)

        results.append({
            "handle": "@" + handle,
            "profile_url": f"https://www.threads.com/@{handle}",
            "title": title,
            "snippet": snippet,
        })
        if len(results) >= max_results:
            break

    cache_set(cache_key, results)
    return results


def multi_query_search(queries, max_total=20):
    """多組 query 合併去重，每次間隔 1.5 秒避免限流"""
    seen = set()
    all_results = []
    for q in queries:
        try:
            found = ddg_search(q, max_results=12)
            for r in found:
                h = r["handle"]
                if h not in seen:
                    seen.add(h)
                    all_results.append(r)
            if len(all_results) >= max_total:
                break
            time.sleep(1.5)
        except Exception:
            continue
    return all_results[:max_total]


@app.route("/")
def index():
    return render_template("index.html")


# ── 批次抓取帳號個人資料 ──
@app.route("/fetch-profiles", methods=["POST"])
def fetch_profiles():
    handles = request.json.get("handles", [])[:15]  # 最多 15 個
    results = {}
    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {pool.submit(fetch_profile, h): h for h in handles}
        for future in as_completed(futures):
            h = futures[future]
            try:
                results[h] = future.result()
            except Exception:
                results[h] = {"handle": h}
    return jsonify(results)


# ── Tab 1：一般帳號查找 ──
@app.route("/find-accounts", methods=["POST"])
def find_accounts():
    data = request.json
    keyword = data.get("keyword", "").strip()
    category = data.get("category", "").strip()

    results = []
    search_ok = False

    parts = [p for p in ["site:threads.com", category, keyword] if p]
    query = " ".join(parts)

    rate_limited = False
    try:
        results = ddg_search(query, max_results=15)
        search_ok = bool(results)
        if not results:
            rate_limited = True
    except Exception:
        rate_limited = True

    tags = CATEGORY_TAGS.get(category, [])
    if keyword:
        tags = [keyword] + tags
    tag_links = [
        {"tag": t, "url": f"https://www.threads.com/tag/{urllib.parse.quote(t)}"}
        for t in tags[:6]
    ]

    return jsonify({"results": results[:12], "tag_links": tag_links,
                    "search_ok": search_ok, "rate_limited": rate_limited})


# ── Tab 1：發文紀錄搜尋（新功能）──
@app.route("/find-by-post", methods=["POST"])
def find_by_post():
    data = request.json
    feature = data.get("feature", "").strip()    # 預設功能選項
    custom_kw = data.get("custom_kw", "").strip()  # 自訂關鍵字

    if custom_kw:
        queries = [
            f'site:threads.com {custom_kw}',
            f'site:threads.com {custom_kw} 推薦 分享',
            f'site:threads.com {custom_kw} app 好用',
        ]
    else:
        base_queries = FEATURE_QUERIES.get(feature, [])
        queries = [f'site:threads.com {q}' for q in base_queries]
        if not queries:
            queries = [f'site:threads.com {feature}']

    rate_limited = False
    try:
        results = multi_query_search(queries, max_total=20)
        if not results:
            rate_limited = True
    except Exception as e:
        rate_limited = True

    kw = custom_kw or feature
    tag_links = [
        {"tag": kw, "url": f"https://www.threads.com/tag/{urllib.parse.quote(kw)}"},
        {"tag": kw + "推薦", "url": f"https://www.threads.com/tag/{urllib.parse.quote(kw+'推薦')}"},
        {"tag": kw + "分享", "url": f"https://www.threads.com/tag/{urllib.parse.quote(kw+'分享')}"},
        {"tag": "美顏相機", "url": f"https://www.threads.com/tag/{urllib.parse.quote('美顏相機')}"},
    ]

    return jsonify({
        "results": results,
        "tag_links": tag_links,
        "search_ok": bool(results),
        "rate_limited": rate_limited,
        "total": len(results),
    })


# ── Tab 2：KOC 評估工具 ──
@app.route("/evaluate", methods=["POST"])
def evaluate():
    data = request.json
    candidates = data.get("candidates", [])
    results = []

    for c in candidates:
        name = c.get("name", "")
        handle = c.get("handle", "")
        followers = int(c.get("followers", 0))
        avg_views = int(c.get("avg_views", 0))
        audience_age = c.get("audience_age", "")
        notes = c.get("notes", "")

        rule1 = followers > 1000
        rule2 = 20000 <= avg_views <= 50000
        rule2_over = avg_views > 50000
        roi = round(avg_views / followers, 1) if followers > 0 else 0

        p = build_price(followers)
        price_range, price_mid = p["range"], p["mid"]

        is_special = followers >= 50000 or avg_views > 50000
        if is_special:
            price_range, price_mid = price_range + " ⭐", price_mid

        all_pass = rule1 and rule2
        is_mutual = not rule1 or (0 < avg_views < 20000)

        results.append({
            "name": name, "handle": handle,
            "followers": followers, "avg_views": avg_views,
            "roi": f"1:{roi}", "rule1": rule1, "rule2": rule2, "rule2_over": rule2_over,
            "audience_age": audience_age, "price_range": price_range, "price_mid": price_mid,
            "is_special": is_special, "notes": notes,
            "type": "qualified" if (all_pass and not is_mutual) else "mutual",
        })

    results.sort(key=lambda x: (0 if x["type"] == "qualified" else 1, -x["followers"]))
    qualified = [r for r in results if r["type"] == "qualified"]
    mutual = [r for r in results if r["type"] == "mutual"]

    return jsonify({
        "results": results,
        "stats": {
            "qualified": len(qualified),
            "mutual": len(mutual),
            "budget": sum(r["price_mid"] for r in qualified if r["price_mid"]),
            "specials": len([r for r in qualified if r["is_special"]]),
        }
    })


def call_worker(cmd, timeout=60):
    """呼叫 threads_worker.py subprocess，傳入 JSON 指令，回傳結果"""
    try:
        result = subprocess.run(
            [sys.executable, WORKER],
            input=json.dumps(cmd),
            capture_output=True, text=True, timeout=timeout
        )
        if result.returncode != 0:
            return {"ok": False, "error": result.stderr[:200] or "Worker 執行失敗"}
        return json.loads(result.stdout)
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": "操作超時，請稍後再試"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def threads_logged_in():
    if not os.path.exists(THREADS_COOKIES):
        return False
    try:
        data = json.load(open(THREADS_COOKIES))
        return bool(data.get("ds_user_id") or data.get("sessionid"))
    except Exception:
        return False


def fetch_threads_posts_api(username, num_posts=5):
    """直接打 Threads 行動版 REST API 拿近期貼文，不需要開瀏覽器"""
    if not os.path.exists(THREADS_COOKIES):
        return None
    try:
        cookies = json.load(open(THREADS_COOKIES))
    except Exception:
        return None
    if not (cookies.get("sessionid") or cookies.get("ds_user_id")):
        return None

    sess = requests.Session()
    sess.cookies.update(cookies)

    try:
        # Step 1：用 username 取得 user_id
        r = sess.get(
            f"https://www.threads.net/api/v1/users/{username}/usernameinfo/",
            headers=THREADS_APP_HEADERS, timeout=10,
        )
        if r.status_code != 200:
            return None
        user_id = r.json().get("user", {}).get("pk")
        if not user_id:
            return None

        # Step 2：取得該用戶的 posts
        r = sess.get(
            f"https://www.threads.net/api/v1/text_feed/{user_id}/profile/",
            headers=THREADS_APP_HEADERS, timeout=10,
        )
        if r.status_code != 200:
            return None

        data = r.json()
        posts = []
        for thread in data.get("threads", []):
            for item in thread.get("thread_items", []):
                post = item.get("post", {})
                view_count = post.get("view_count")
                like_count = post.get("like_count")
                if like_count is not None or view_count is not None:
                    posts.append({
                        "views": int(view_count) if view_count is not None else None,
                        "likes": int(like_count) if like_count is not None else None,
                        "replies": post.get("direct_reply_count"),
                    })
                    if len(posts) >= num_posts:
                        break
            if len(posts) >= num_posts:
                break

        return posts if posts else None
    except Exception:
        return None


def build_price(followers):
    """Threads 圖文定價"""
    if followers is None:
        return {"range": "待確認", "mid": None, "tier": "unknown", "is_mutual": False}
    if followers < 1000:
        return {"range": "互惠", "mid": None, "tier": "micro", "is_mutual": True}
    if followers < 10000:
        return {"range": "互惠～NT$1,500", "mid": 750, "tier": "nano", "is_mutual": False}
    if followers < 15000:
        return {"range": "NT$1,500～3,500", "mid": 2500, "tier": "mini", "is_mutual": False}
    if followers < 50000:
        return {"range": "NT$3,500～15,000", "mid": 9250, "tier": "small", "is_mutual": False}
    if followers < 100000:
        return {"range": "NT$15,000～35,000", "mid": 25000, "tier": "mid", "is_mutual": False}
    return {"range": "NT$35,000 以上", "mid": None, "tier": "macro", "is_mutual": False}


def build_price_ig(followers):
    """IG 定價：Reels / 限動 / 三合一"""
    if followers is None:
        return {"reels": "待確認", "story": "待確認", "combo": "待確認",
                "reels_mid": None, "story_mid": None, "combo_mid": None, "is_mutual": False}
    if followers < 1000:
        return {"reels": "互惠", "story": "互惠", "combo": "互惠",
                "reels_mid": None, "story_mid": None, "combo_mid": None, "is_mutual": True}
    if followers < 10000:
        return {"reels": "NT$3,500", "story": "互惠～NT$1,500", "combo": "互惠～NT$6,000",
                "reels_mid": 3500, "story_mid": 750, "combo_mid": 3000, "is_mutual": False}
    if followers < 15000:
        return {"reels": "NT$4,500～8,000", "story": "NT$1,500～3,000", "combo": "NT$6,000～12,000",
                "reels_mid": 6250, "story_mid": 2250, "combo_mid": 9000, "is_mutual": False}
    if followers < 50000:
        return {"reels": "NT$8,000～20,000", "story": "NT$3,500～20,000", "combo": "NT$15,000～25,000",
                "reels_mid": 14000, "story_mid": 11750, "combo_mid": 20000, "is_mutual": False}
    if followers < 100000:
        return {"reels": "NT$25,000～55,000", "story": "NT$25,000～35,000", "combo": "NT$30,000～65,000",
                "reels_mid": 40000, "story_mid": 30000, "combo_mid": 47500, "is_mutual": False}
    return {"reels": "NT$55,000 以上", "story": "NT$35,000 以上", "combo": "NT$65,000 以上",
            "reels_mid": None, "story_mid": None, "combo_mid": None, "is_mutual": False}


def generate_queries(topic, style):
    """根據推廣主題和KOC風格，產生多組 Threads 搜尋 query"""
    # 拆出關鍵詞
    topic_kw = topic.strip()
    style_kw = re.sub(r'[、,，\s]+', ' ', style).strip()
    style_parts = re.split(r'[、,，\s]+', style)[:3]

    queries = []
    # 主題 + 風格
    if topic_kw and style_kw:
        queries.append(f'site:threads.com {topic_kw} {style_parts[0] if style_parts else ""}')
    # 主題 + 推薦/分享
    if topic_kw:
        queries.append(f'site:threads.com {topic_kw} 推薦 分享')
        queries.append(f'site:threads.com {topic_kw} 修圖 自拍')
    # 風格關鍵詞
    for kw in style_parts[:2]:
        if kw:
            queries.append(f'site:threads.com {kw} 美妝 自拍')
    # 美顏相機相關
    if topic_kw:
        queries.append(f'site:threads.com 美顏相機 {topic_kw}')
    return [q for q in queries if q.strip()][:6]


def ig_hashtag_to_queries(topic, style):
    """從推廣主題和風格，產生要搜尋的 IG hashtag 列表"""
    tags = []
    # 主題關鍵字轉 hashtag
    for kw in re.split(r'[、,，\s]+', topic):
        kw = kw.strip()
        if kw:
            tags.append(kw)
            tags.append(kw + "推薦")
    # 風格關鍵字
    for kw in re.split(r'[、,，\s]+', style)[:3]:
        kw = kw.strip()
        if kw:
            tags.append(kw)
    # 固定加美顏相機相關
    tags += ["美顏相機", "美妝分享", "修圖推薦"]
    # 去重、限制數量
    seen = set()
    result = []
    for t in tags:
        if t not in seen:
            seen.add(t)
            result.append(t)
    return result[:6]


@app.route("/smart-search", methods=["POST"])
def smart_search():
    data = request.json
    topic = data.get("topic", "").strip()
    style = data.get("style", "").strip()
    if not topic and not style:
        return jsonify({"ok": False, "error": "請填入推廣主題或 KOC 風格"})

    if not threads_logged_in():
        return jsonify({"ok": False, "error": "請先完成 Threads 登入（點上方「連線 Threads」按鈕）"})

    # 產生更精準的搜尋關鍵字：主題片語 + 風格詞組合
    style_parts = [s.strip() for s in re.split(r'[、,，\s]+', style) if s.strip()]
    topic_parts = [t.strip() for t in re.split(r'[、,，\s]+', topic) if t.strip()]

    keywords = []
    # 風格＋主題組合句（最精準）
    if style_parts and topic_parts:
        keywords.append(f"{style_parts[0]} {topic_parts[0]}")
    # 主題＋常見分享詞
    if topic_parts:
        keywords.append(f"{topic_parts[0]} 分享")
        keywords.append(f"{topic_parts[0]} 推薦")
    # 純風格詞
    for s in style_parts[:2]:
        keywords.append(s)
    # 去重保留順序
    seen_kw = set()
    final_keywords = []
    for kw in keywords:
        if kw not in seen_kw:
            seen_kw.add(kw)
            final_keywords.append(kw)
    keywords = final_keywords[:6]

    worker_result = call_worker({"action": "search", "keywords": keywords, "cookies_path": THREADS_COOKIES}, timeout=90)
    if not worker_result.get("ok"):
        return jsonify({"ok": False, "error": worker_result.get("error", "搜尋失敗")})
    raw = worker_result.get("results", [])
    if not raw:
        return jsonify({"ok": True, "results": [], "mutual": [], "keywords": keywords, "msg": "未找到帳號，請調整關鍵字後重試"})

    # 批次抓 Threads profile（用 OG tags，不需登入）
    handles = [r["handle"] for r in raw]
    profiles = {}
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(fetch_profile, h): h for h in handles}
        for future in as_completed(futures):
            h = futures[future]
            try:
                profiles[h] = future.result()
            except Exception:
                profiles[h] = {"handle": h}

    results = []   # 付費合作（1K–10萬粉）
    mutual = []    # 互惠候選（<1K 或粉絲數不明）

    for r in raw:
        h = r["handle"]
        p = profiles.get(h, {})
        followers = p.get("followers")

        # 過濾超過 10 萬粉的帳號
        if followers is not None and followers > 100000:
            continue

        price = build_price(followers)
        card = {
            "handle": h,
            "display_name": p.get("display_name") or h,
            "avatar": p.get("avatar", ""),
            "bio": p.get("bio", ""),
            "followers": followers,
            "posts": p.get("posts"),
            "snippet": r.get("snippet", ""),
            "price_range": price["range"],
            "price_mid": price["mid"],
            "price_tier": price["tier"],
            "profile_url": r.get("profile_url", f"https://www.threads.com/{h}"),
        }

        # 互惠候選：粉絲數不明 或 < 1000
        if followers is None or followers < 1000:
            mutual.append(card)
        else:
            results.append(card)

    results.sort(key=lambda x: -(x["followers"] or 0))
    mutual.sort(key=lambda x: -(x["followers"] or 0))
    return jsonify({"ok": True, "results": results, "mutual": mutual, "keywords": keywords})


# ── Instagram 登入 ──
@app.route("/ig-login", methods=["POST"])
def ig_login():
    global _ig, _ig_username, _ig_password, _ig_pending
    data = request.json
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    if not username or not password:
        return jsonify({"ok": False, "error": "請填入帳號和密碼"})
    try:
        loader = instaloader.Instaloader(
            download_pictures=False, download_videos=False,
            download_video_thumbnails=False, download_geotags=False,
            download_comments=False, save_metadata=False, quiet=True,
        )
        loader.login(username, password)
        _ig = loader
        _ig_username = username
        _ig_password = password
        try:
            loader.save_session_to_file(IG_SESSION_FILE)
            open(IG_USERNAME_FILE, "w").write(username)
        except Exception:
            pass
        return jsonify({"ok": True, "username": username})
    except instaloader.exceptions.BadCredentialsException:
        return jsonify({"ok": False, "error": "帳號或密碼錯誤"})
    except instaloader.exceptions.TwoFactorAuthRequiredException:
        _ig_pending = (loader, username)
        return jsonify({"ok": False, "need_2fa": True, "error": "需要雙重驗證，請輸入收到的驗證碼"})
    except instaloader.exceptions.LoginException:
        return jsonify({"ok": False, "error": "登入失敗，請確認帳號密碼是否正確"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ig-verify", methods=["POST"])
def ig_verify():
    global _ig, _ig_username, _ig_pending
    if not _ig_pending:
        return jsonify({"ok": False, "error": "沒有待驗證的登入，請重新登入"})
    loader, username = _ig_pending
    code = request.json.get("code", "").strip().replace(" ", "")
    if not code:
        return jsonify({"ok": False, "error": "請輸入驗證碼"})
    try:
        loader.two_factor_login(code)
        _ig = loader
        _ig_username = username
        _ig_pending = None
        try:
            loader.save_session_to_file(IG_SESSION_FILE)
            open(IG_USERNAME_FILE, "w").write(username)
        except Exception:
            pass
        return jsonify({"ok": True, "username": username})
    except instaloader.exceptions.BadCredentialsException:
        return jsonify({"ok": False, "error": "驗證碼錯誤，請重新輸入"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/threads-login", methods=["POST"])
def do_threads_login():
    """接收前端貼入的 Cookie-Editor JSON，解析後儲存為 .threads_cookies.json"""
    data = request.json or {}
    raw = data.get("cookies_json", "").strip()
    if not raw:
        return jsonify({"ok": False, "error": "請貼入 cookies JSON"})
    try:
        items = json.loads(raw)
        # 支援 Cookie-Editor 匯出格式（list of {name, value, ...}）
        # 也支援直接傳 dict {name: value, ...}
        if isinstance(items, list):
            cookie_dict = {c["name"]: c["value"] for c in items if "name" in c and "value" in c}
        elif isinstance(items, dict):
            cookie_dict = items
        else:
            return jsonify({"ok": False, "error": "格式不正確，請使用 Cookie-Editor 的 JSON 匯出"})
    except Exception as e:
        return jsonify({"ok": False, "error": f"JSON 解析失敗：{e}"})

    if not (cookie_dict.get("ds_user_id") or cookie_dict.get("sessionid")):
        return jsonify({"ok": False, "error": "未找到有效的 Threads session（ds_user_id / sessionid），請確認在 threads.com 登入後再匯出"})

    with open(THREADS_COOKIES, "w") as f:
        json.dump(cookie_dict, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True, "message": f"已儲存 {len(cookie_dict)} 個 cookies"})


@app.route("/threads-status", methods=["GET"])
def threads_status():
    return jsonify({"logged_in": threads_logged_in()})


@app.route("/ig-status", methods=["GET"])
def ig_status():
    return jsonify({"logged_in": _ig is not None, "username": _ig_username})


@app.route("/ig-fetch-profile", methods=["POST"])
def ig_fetch_profile():
    if not _ig:
        return jsonify({"ok": False, "error": "請先登入 Instagram"})
    data = request.json
    username = data.get("username", "").lstrip("@").strip()
    cache_key = "ig_" + username
    cached = cache_get(cache_key)
    if cached:
        return jsonify({"ok": True, **cached})
    try:
        profile = instaloader.Profile.from_username(_ig.context, username)
        result = {
            "handle": "@" + username,
            "display_name": profile.full_name or username,
            "avatar": profile.profile_pic_url,
            "bio": (profile.biography or "")[:100],
            "followers": profile.followers,
            "posts": profile.mediacount,
            "platform": "instagram",
            "profile_url": f"https://www.instagram.com/{username}/",
        }
        cache_set(cache_key, result)
        return jsonify({"ok": True, **result})
    except instaloader.exceptions.ProfileNotExistsException:
        return jsonify({"ok": False, "error": f"找不到帳號 @{username}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/export-excel", methods=["POST"])
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    import io

    data = request.json or {}
    kocs = data.get("kocs", [])
    topic = data.get("topic", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KOC名單"

    headers = ["主題", "负责人", "狀態", "發佈日期", "發文平臺",
               "網紅名字", "網紅帳號", "Threads\n帳號連結", "IG\n帳號連結",
               "粉絲數", "合作價格", "備註"]
    col_widths = [14, 10, 10, 12, 18, 14, 18, 40, 36, 10, 12, 30]

    header_fill = PatternFill("solid", fgColor="FFD6E8")
    header_font = Font(bold=True, size=11)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 32

    for i, koc in enumerate(kocs, 2):
        handle_clean = koc.get("handle", "").lstrip("@")
        is_mutual = koc.get("is_mutual", False)
        followers = koc.get("followers")
        if is_mutual:
            price_val = "互惠"
        else:
            price_val = koc.get("price_mid") or koc.get("price_range", "")

        row_data = [
            topic or koc.get("topic", ""),
            "peggy",
            "待過審",
            None,
            "Threads ＋限動",
            koc.get("display_name", ""),
            handle_clean,
            koc.get("profile_url", f"https://www.threads.com/@{handle_clean}"),
            "",
            followers,
            price_val,
            koc.get("bio", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=col in (8, 9, 12))
            cell.border = thin
        ws.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"KOC名單_{time.strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/quick-eval", methods=["POST"])
def quick_eval():
    data = request.json or {}
    username = data.get("username", "").lstrip("@").strip()
    platform = data.get("platform", "threads")

    if not username:
        return jsonify({"ok": False, "error": "請輸入帳號名稱"})

    cache_key = f"qe_{platform}_{username}"
    cached = cache_get(cache_key)
    if cached:
        return jsonify({"ok": True, **cached})

    if platform == "instagram":
        if not _ig:
            return jsonify({"ok": False, "error": "請先登入 Instagram（點上方「登入 Instagram」按鈕）"})
        try:
            profile = instaloader.Profile.from_username(_ig.context, username)
        except instaloader.exceptions.ProfileNotExistsException:
            return jsonify({"ok": False, "error": f"找不到帳號 @{username}"})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)})

        recent_posts = []
        posts_error = None
        try:
            it = iter(profile.get_posts())
            for i in range(5):
                try:
                    post = next(it)
                except StopIteration:
                    break
                except Exception as pe:
                    posts_error = str(pe)
                    break
                views = post.video_views if post.is_video else None
                recent_posts.append({
                    "index": i + 1,
                    "views": views,
                    "likes": post.likes,
                    "is_video": post.is_video,
                    "url": f"https://www.instagram.com/p/{post.shortcode}/",
                })
        except Exception as e:
            posts_error = str(e)

        # 用 mobile API 拿完整 profile（不打 GraphQL，不被限流）
        full_name = profile._node.get("full_name") or username
        avatar = (profile._node.get("hd_profile_pic_url_info") or {}).get("url") or profile._node.get("profile_pic_url") or ""
        followers, bio, posts_count = None, "", None

        try:
            ig_sess = _ig.context._session
            ig_headers = {
                "User-Agent": "Instagram 269.0.0.18.75 Android",
                "X-IG-App-ID": "936619743392459",
                "Accept-Language": "zh-TW",
            }
            r = ig_sess.get(
                f"https://i.instagram.com/api/v1/users/web_profile_info/?username={username}",
                headers=ig_headers, timeout=10,
            )
            if r.status_code == 200:
                udata = r.json().get("data", {}).get("user", {})
                followers  = udata.get("edge_followed_by", {}).get("count")
                posts_count = udata.get("edge_owner_to_timeline_media", {}).get("count")
                bio        = (udata.get("biography") or "")[:100]
                full_name  = udata.get("full_name") or full_name
                avatar     = udata.get("profile_pic_url_hd") or udata.get("profile_pic_url") or avatar
        except Exception:
            pass

        ig_price = build_price_ig(followers)
        view_nums = [p["views"] for p in recent_posts if p["views"] is not None]
        like_nums = [p["likes"] for p in recent_posts if p["likes"] is not None]
        avg_views = int(sum(view_nums) / len(view_nums)) if view_nums else None
        avg_likes = int(sum(like_nums) / len(like_nums)) if like_nums else None

        result = {
            "handle": "@" + username,
            "display_name": full_name,
            "avatar": avatar,
            "bio": bio,
            "followers": followers,
            "posts_count": posts_count,
            "recent_posts": recent_posts,
            "avg_views": avg_views,
            "avg_likes": avg_likes,
            "ig_price": ig_price,
            "price_range": ig_price["reels"],
            "posts_error": posts_error,
            "platform": "instagram",
            "profile_url": f"https://www.instagram.com/{username}/",
        }
        cache_set(cache_key, result)
        return jsonify({"ok": True, **result})

    else:  # threads
        profile = fetch_profile("@" + username)
        followers = profile.get("followers")
        price = build_price(followers)
        recent_posts = []

        if threads_logged_in():
            # 優先：直接打行動版 API（快、穩，約 1-2 秒）
            api_posts = fetch_threads_posts_api(username, num_posts=5)
            if api_posts is not None:
                recent_posts = api_posts
            else:
                # fallback：Playwright 攔截網路請求（約 15-20 秒）
                wr = call_worker(
                    {"action": "profile_posts", "username": username, "cookies_path": THREADS_COOKIES},
                    timeout=55,
                )
                if wr.get("ok"):
                    recent_posts = wr.get("posts", [])

        # Threads API 只回 like_count（view_count 是作者才能看的數據）
        for i, p in enumerate(recent_posts):
            p.setdefault("index", i + 1)
        like_nums = [p["likes"] for p in recent_posts if p.get("likes") is not None]
        avg_likes = int(sum(like_nums) / len(like_nums)) if like_nums else None

        result = {
            "handle": "@" + username,
            "display_name": profile.get("display_name", username),
            "avatar": profile.get("avatar", ""),
            "bio": profile.get("bio", ""),
            "followers": followers,
            "posts_count": profile.get("posts"),
            "recent_posts": recent_posts,
            "avg_views": None,
            "avg_likes": avg_likes,
            "price_range": price["range"],
            "price_mid": price["mid"],
            "price_tier": price["tier"],
            "platform": "threads",
            "profile_url": f"https://www.threads.com/@{username}",
        }
        cache_set(cache_key, result)
        return jsonify({"ok": True, **result})


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False, port=5001)
